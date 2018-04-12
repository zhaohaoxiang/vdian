#!/usr/local/bin/python3.6

##
## Description:
## 	Parse Excels from vdian
##
## Author: Harry Zhao


import re
import os
import sys
import shutil
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from product import *


# Use nargs to specify how many arguments an option should take.
ap = argparse.ArgumentParser()
ap.add_argument('--sum', dest="is_sum", help='show summary', action='store_true')
ap.set_defaults(is_sum=False)
ap.add_argument('--obj', dest="obj", help='show target')
ap.add_argument('--file', dest="excel_file", help='parse file')
args = ap.parse_args()
	
	
##
## Show usage
##
def usage():
	print("Usage: ./vdian.py <file>\n")
	

##
## Parse workbook
##	U: description
##	L: buy number
##
def parse(file, vdian, a_vdian):
	wb = load_workbook(file)
	#sheet = wb['ALL']
	sheet = wb.active
	row_count = sheet.max_row
	#print(sheet['A2'].value)
	
	for i in range(2, row_count + 1):
	
		# K: description
		ii = 0
		while (sheet['K' + str(i-ii)].value is None):
			ii = ii + 1
		p_name = sheet['K' + str(i-ii)].value
		if not (sheet['L' + str(i)].value is None):
			p_name = p_name + sheet['L' + str(i)].value
		
		# O: buy number
		num = sheet['O' + str(i)].value
		
		# A: order ID
		ii = 0
		while (sheet['A' + str(i-ii)].value is None):
			ii = ii + 1
		id = sheet['A' + str(i-ii)].value
		
		# H: pay date
		ii = 0
		while (sheet['H' + str(i-ii)].value is None):
			ii = ii + 1
		pay_date = sheet['H' + str(i-ii)].value
		
		# X: name
		ii = 0
		while (sheet['X' + str(i-ii)].value is None):
			ii = ii + 1
		name = sheet['X' + str(i-ii)].value
		
		# Y: phone
		ii = 0
		while (sheet['Y' + str(i-ii)].value is None):
			ii = ii + 1
		phone = sheet['Y' + str(i-ii)].value
		
		# P: price
		price = sheet['P' + str(i)].value
		
		# S: refund status
		refund = sheet['S' + str(i)].value
				
		# Z: province
		ii = 0
		while (sheet['Z' + str(i-ii)].value is None):
			ii = ii + 1
		province = sheet['Z' + str(i-ii)].value
		
		# AA: city
		ii = 0
		while (sheet['AA' + str(i-ii)].value is None):
			ii = ii + 1
		city = sheet['AA' + str(i-ii)].value
		
		# AB: district
		ii = 0
		while (sheet['AB' + str(i-ii)].value is None):
			ii = ii + 1
		district = sheet['AB' + str(i-ii)].value
		
		# AC: address
		ii = 0
		while (sheet['AC' + str(i-ii)].value is None):
			ii = ii + 1
		addr = sheet['AC' + str(i-ii)].value
		
		# E: status
		ii = 0
		while (sheet['E' + str(i-ii)].value is None):
			ii = ii + 1
		status = sheet['E' + str(i-ii)].value

		if (not ('待发货' in status)) and (not args.is_sum):
			print("[WARNING] " + status + " : " + id + " : " + name + " : " + phone)
			continue
		if args.is_sum and (not ('已发货' in status or '已完成' in status)):
			print("[WARNING] " + status + " : " + id + " : " + name + " : " + phone)
			continue
		
		# Z: summary 1
		# X: summary 2
		ii = 0
		while (sheet['E' + str(i-ii)].value is None):
			ii = ii + 1
		note_1 = sheet['AQ' + str(i-ii)].value
		ii = 0
		while (sheet['E' + str(i-ii)].value is None):
			ii = ii + 1
		note_2 = sheet['AO' + str(i-ii)].value
		if note_1 is None:
			note_1 = ""
		if note_2 is None:
			note_2 = ""
		note = note_1 + ' ' + note_2
		
		if ("退款" in refund) and (not "退款关闭" in refund):
			# T: refund money
			refund_money = sheet['T' + str(i)].value

			# B: total money
			ii = 0
			while (sheet['E' + str(i-ii)].value is None):
				ii = ii + 1
			total_money = sheet['B' + str(i-ii)].value
			
			start = i - ii
			ii = 1
			sum_money = float(sheet['P' + str(start)].value) * int(num)
			#print("[DBG] sum_money " + str(sum_money) + " : num " + str(num) + " : total " + str(sheet['B' + str(start)].value) + " : sum_money " + str(sheet['P' + str(start)].value))
			while (sheet['E' + str(start+ii)].value is None):
				sum_money = float(sum_money) + float(sheet['P' + str(start+ii)].value) * int(float(sheet['O' + str(start+ii)].value))
				ii = ii + 1
				#print("[DBG] sum_money " + str(sum_money))
			#print("[DBG] sum_money " + str(sum_money))
			sum_money = float(sum_money) - float(sheet['P' + str(i)].value) * int(num)
			sum_money = float(total_money) - float(sum_money)

			print("[REFUND] " + status + " : " + id + " : " + name + " : " + phone + " : " + str(total_money) + " - " + str(sum_money) + " - " + str(refund_money))
			if (sum_money == refund_money):
				continue
			else:
				actual_num = (float(sum_money) - float(refund_money)) / float(price)
				#actual_num = int(actual_num)
				print("[REFUND] Adjust num from " + str(num) + " to " + str(actual_num))
				num = actual_num
				if (num == 0):
					continue
		
		if (args.obj is None) or ((not (args.obj is None)) and (args.obj in p_name)):
			vdian.add_order(p_name, price, id, name, phone, num, province, city, district, addr, note, pay_date)
			a_vdian.add_order(p_name, price, id, name, phone, num, province, city, district, addr, note, pay_date)


def create(file, vdian):
	for p in vdian.products:
		wb = Workbook()
		#ws = wb.create_sheet("sheet")
		ws = wb.active
		
		# Summary
		ws['A1'] = "Product"
		ws['A2'] = p.p_name
		ws['B1'] = "Price"
		ws['B2'] = p.price
		ws['C1'] = "Order Number"
		ws['C2'] = p.order_num
		ws['D1'] = "Sell Number"
		ws['D2'] = p.sell_num
		
		# Details
		ws['A4'] = "ID"
		ws['B4'] = 'Name'
		ws['C4'] = 'Date'
		ws['D4'] = 'Phone'
		#ws['E1'] = 'Price'
		ws['E4'] = 'Number'
		ws['F4'] = 'Province'
		ws['G4'] = 'City'
		ws['H4'] = 'District'
		ws['I4'] = 'Address'
		ws['J4'] = 'Note'
		ws['K4'] = 'Product'
		for j in range(0, p.order_num):
			o = p.orders[j]
			i = j + 5
			ws['A' + str(i)] = o.id
			ws['B' + str(i)] = o.name
			ws['C' + str(i)] = o.pay_date
			ws['D' + str(i)] = o.phone
			#ws['E' + str(i)] = p.price
			ws['E' + str(i)] = o.num
			ws['F' + str(i)] = o.province
			ws['G' + str(i)] = o.city
			ws['H' + str(i)] = o.district
			ws['I' + str(i)] = o.addr
			ws['J' + str(i)] = o.note
			ws['K' + str(i)] = p.p_name
		#new_file = p.p_name[0:10] + ".xlsx"
		#new_file = "output/" + p.p_name[0:10] + "-" + file
		new_file = "output/" + p.p_name + "-" + file
		wb.save(new_file)
				

##
## Process
##
def process(file, a_vdian):
	vdian = Vdian()
	parse(file, vdian, a_vdian)
	create(file, vdian)


##
## Main
##
def main():
	if not (args.obj is None):
		print("Parsing " + args.obj)
		
	out = "output"
	if os.path.exists(out):
		shutil.rmtree(out)
	os.makedirs(out)

	a_vdian = Vdian()

	file_num = 0
	
	if args.excel_file is None:
		files = os.listdir('.')
		for file in files:
			if file.lower().endswith('.xls') and not file.lower().startswith('~'):
				print("Please convert " + file + " into xlsx")
			if file.lower().endswith('.xlsx') and not file.lower().startswith('~'):
				file_num = file_num + 1
				process(file, a_vdian)
				os.remove(file)
	else:
		file = args.excel_file
		process(file, a_vdian)
		
	a_vdian.show()
	if file_num > 1:
		create("summary.xlsx", a_vdian)

		
##
## Starting ...
##
main()
