#!/usr/local/bin/python3.6

##
## Description:
## 	Parse Excels from vdian
##
## Author: Harry Zhao


import re
import os
import sys
import shutil
import argparse
from openpyxl import Workbook
from openpyxl import load_workbook
from product import *


# Use nargs to specify how many arguments an option should take.
ap = argparse.ArgumentParser()
ap.add_argument('--sum', dest="is_sum", help='show summary', action='store_true')
ap.set_defaults(is_sum=False)
ap.add_argument('--file', dest="excel_file", help='parse file')
args = ap.parse_args()
	
	
##
## Show usage
##
def usage():
	print("Usage: ./vdian.py <file>\n")
	

##
## Parse workbook
##	U: description
##	L: buy number
##
def parse(file, vdian, a_vdian):
	wb = load_workbook(file)
	#sheet = wb['ALL']
	sheet = wb.active
	row_count = sheet.max_row
	#print(sheet['A2'].value)
	
	for i in range(2, row_count + 1):
	
		# U: description
		p_names = sheet['U' + str(i)].value
		found = False
		p_names_temp = ''
		for j in range(0, len(p_names)):
			if p_names[j] == '[':
				found = True
			if not found:
				p_names_temp = p_names_temp + p_names[j]
			if p_names[j] == ']':
				found = False
		p_names = p_names_temp
		p_name_a = p_names.split(';')
		#new_name = ''
		#for name in p_name_a:
		#	pos = name.find('[')
		#	new_name = new_name + name[0:pos] + ';'
		#new_name = new_name[0:-1]
		#p_name_a = new_name.split(';')
		
		#p_names = re.sub('['+s+':\d+]', '', p_names)
		#p_names = re.sub('[\w]', '', p_names)
		#p_name_a = p_names.split(';')
		
		# L: buy number
		nums = sheet['L' + str(i)].value
		num_a = nums.split(';')
		
		# A: order ID
		id = sheet['A' + str(i)].value
		
		# F: pay date
		pay_date = sheet['F' + str(i)].value
		
		# I: name
		name = sheet['I' + str(i)].value
		
		# J: phone
		phone = sheet['J' + str(i)].value
		
		# M: price
		prices = sheet['M' + str(i)].value
		price_a = prices.split(';')
		
		# N: province
		province = sheet['N' + str(i)].value
		
		# O: city
		city = sheet['O' + str(i)].value
		
		# P: district
		district = sheet['P' + str(i)].value
		
		# Q: address
		addr = sheet['Q' + str(i)].value
		
		# C: status
		status = sheet['C' + str(i)].value
		if (not ('待发货' in status)) and (not args.is_sum):
			print("[SKIP] " + status + " : " + id + " : " + name + " : " + phone)
			continue
		if args.is_sum and (not ('待发货' in status or '已发货' in status or '已完成' in status)):
			print("[SKIP] " + status + " : " + id + " : " + name + " : " + phone)
			continue
		
		# Z: summary 1
		# AB: summary 2
		note_1 = sheet['Z' + str(i)].value
		note_2 = sheet['AB' + str(i)].value
		note = note_1 + ' ' + note_2
		
		for j in range(0, len(p_name_a)):
			p_name = p_name_a[j]
			p_name = p_name.strip()
			num = num_a[j]
			price = price_a[j]
			vdian.add_order(p_name, price, id, name, phone, num, province, city, district, addr, note, pay_date)
			a_vdian.add_order(p_name, price, id, name, phone, num, province, city, district, addr, note, pay_date)


def create(file, vdian):
	for p in vdian.products:
		wb = Workbook()
		#ws = wb.create_sheet("sheet")
		ws = wb.active
		
		# Summary
		ws['A1'] = "Product"
		ws['A2'] = p.p_name
		ws['B1'] = "Price"
		ws['B2'] = p.price
		ws['C1'] = "Order Number"
		ws['C2'] = p.order_num
		ws['D1'] = "Sell Number"
		ws['D2'] = p.sell_num
		
		# Details
		ws['A4'] = "ID"
		ws['B4'] = 'Name'
		ws['C4'] = 'Date'
		ws['D4'] = 'Phone'
		#ws['E1'] = 'Price'
		ws['E4'] = 'Number'
		ws['F4'] = 'Province'
		ws['G4'] = 'City'
		ws['H4'] = 'District'
		ws['I4'] = 'Address'
		ws['J4'] = 'Note'
		ws['K4'] = 'Product'
		for j in range(0, p.order_num):
			o = p.orders[j]
			i = j + 5
			ws['A' + str(i)] = o.id
			ws['B' + str(i)] = o.name
			ws['C' + str(i)] = o.pay_date
			ws['D' + str(i)] = o.phone
			#ws['E' + str(i)] = p.price
			ws['E' + str(i)] = o.num
			ws['F' + str(i)] = o.province
			ws['G' + str(i)] = o.city
			ws['H' + str(i)] = o.district
			ws['I' + str(i)] = o.addr
			ws['J' + str(i)] = o.note
			ws['K' + str(i)] = p.p_name
		#new_file = p.p_name[0:10] + ".xlsx"
		#new_file = "output/" + p.p_name[0:10] + "-" + file
		new_file = "output/" + p.p_name + "-" + file
		wb.save(new_file)
				

##
## Process
##
def process(file, a_vdian):
	vdian = Vdian()
	parse(file, vdian, a_vdian)
	create(file, vdian)


##
## Main
##
def main():
	out = "output"
	if os.path.exists(out):
		shutil.rmtree(out)
	os.makedirs(out)

	a_vdian = Vdian()

	file_num = 0
	
	if args.excel_file is None:
		files = os.listdir('.')
		for file in files:
			if file.lower().endswith('.xls') and not file.lower().startswith('~'):
				print("Please convert " + file + " into xlsx")
			if file.lower().endswith('.xlsx') and not file.lower().startswith('~'):
				file_num = file_num + 1
				process(file, a_vdian)
	else:
		file = args.excel_file
		process(file, a_vdian)
		
	a_vdian.show()
	if file_num > 1:
		create("summary.xlsx", a_vdian)

		
##
## Starting ...
##
main()
